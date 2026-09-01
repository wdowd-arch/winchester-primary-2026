#!/usr/bin/env python3
"""
Build the reporter data-entry workbook for the Sept. 1, 2026 Winchester count.

The Results tab is the single source of truth for the live page: it is published
as CSV and read every 15 seconds. The column contract must stay in sync with
RACES[] / COLUMNS in ../index.html.

    Column A   section - must EXACTLY match a RACES[].key in index.html
    Column B   item    - candidate name, "Write-in", or "Blanks"
    Columns C+ areas   - the 16 reporting areas, in COLUMNS order: the other
                         communities one column each, Winchester P1-P8

Winchester's eight precincts are split between the two districts, so each race
uses only four of the eight columns. The cells that do not belong to a race are
shaded and locked, because a number typed into the wrong column is the single
easiest mistake to make on election night.

Writes to data/ at the repo root. Run from anywhere:

    python3 tools/build-sheet.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Must match RACES / COLUMNS in ../index.html ───────────────────────────────
#
# Statutory district composition (2021 redistricting, effective 2022):
#   2nd Middlesex - all of Medford, all of Somerville, Winchester precincts 4-7,
#                   and the Cambridge portion: Wards 10 and 11 entire plus Ward 7
#                   Precinct 1 and Ward 8 Precinct 1. Cambridge Ward 9 is NOT in
#                   this district (it is Suffolk and Middlesex). Confirmed against
#                   the Cambridge Election Commission ward/precinct/senate-district
#                   map, effective 2021-12-31.
#   5th Middlesex - all of Malden, Melrose, Reading, Stoneham and Wakefield,
#                   and Winchester precincts 1, 2, 3 and 8.
#
# Winchester is broken out precinct by precinct because that is where our own
# reporters are. Every other community is one column, entered from that
# community's own posted totals.
RACES = [
    {
        "key": "Second Middlesex Senate - Democratic",
        "title": "State Senate - 2nd Middlesex (Democratic primary)",
        "open": "Open seat - Sen. Patricia Jehlen is not seeking re-election.",
        "communities": [
            ("Somerville", ["Somerville"]),
            ("Medford", ["Medford"]),
            ("Cambridge (Wards 10, 11 + 7-1, 8-1)",
             ["Cambridge W10 W11 W7-1 W8-1"]),
            ("Winchester (precincts 4-7)",
             ["Winchester P4", "Winchester P5", "Winchester P6", "Winchester P7"]),
        ],
        "cands": [
            ("Burhan Azeem", "Cambridge"),
            ("Christine Barber", "Somerville"),
            ("Tom Hopcroft", "Winchester"),
            ("Matt McLaughlin", "Somerville"),
            ("Erika Uyterhoeven", "Somerville"),
        ],
    },
    {
        "key": "Fifth Middlesex Senate - Democratic",
        "title": "State Senate - 5th Middlesex (Democratic primary)",
        "open": "Open seat - Sen. Jason Lewis is not seeking re-election.",
        "communities": [
            ("Malden", ["Malden"]),
            ("Melrose", ["Melrose"]),
            ("Reading", ["Reading"]),
            ("Stoneham", ["Stoneham"]),
            ("Wakefield", ["Wakefield"]),
            ("Winchester (precincts 1, 2, 3, 8)",
             ["Winchester P1", "Winchester P2", "Winchester P3", "Winchester P8"]),
        ],
        "cands": [
            ("Kate Lipper-Garabedian", "Melrose"),
            ("Carey McDonald", "Malden"),
            ("Ryan J. O'Malley", "Malden"),
        ],
    },
]

# Sheet columns C onward, in this exact order. Must equal COLUMNS in index.html.
COLUMNS = [a for r in RACES for _, areas in r["communities"] for a in areas]

# Precinct/subprecinct reporting units per community, supplied by the editor
# 2026-08-31, including the "A" subprecincts that report separately at state
# elections. Cambridge is deliberately absent - see the note on that tab.
PRECINCTS = {
    "Somerville": 32, "Medford": 18, "Cambridge (Wards 10, 11 + 7-1, 8-1)": 11, "Winchester (precincts 4-7)": 4,
    "Malden": 27, "Melrose": 14, "Reading": 8, "Stoneham": 7, "Wakefield": 7,
    "Winchester (precincts 1, 2, 3, 8)": 4,
}

INK        = "1A1C1E"
ACCENT     = "35714F"
HEAD_FILL  = PatternFill("solid", fgColor=ACCENT)
TITLE_FILL = PatternFill("solid", fgColor="E9F2EC")
NA_FILL    = PatternFill("solid", fgColor="D9D6D0")   # column not used by this race
ENTRY_FILL = PatternFill("solid", fgColor="FFFFFF")
SUB_FILL   = PatternFill("solid", fgColor="F4F2EE")
THIN       = Side(style="thin", color="BFBBB4")
BOX        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# RESULTS  (this tab is what gets published as CSV)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Results"

header = ["Section", "Item"] + COLUMNS
ws.append(header)
for c in range(1, len(header) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BOX
ws.row_dimensions[1].height = 42
ws.freeze_panes = "C2"

dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dv.error = "Enter a whole number of votes, or leave blank if this community has not reported."
dv.errorTitle = "Votes must be a whole number"
ws.add_data_validation(dv)

col_of = {name: 3 + i for i, name in enumerate(COLUMNS)}

row = 2
for race in RACES:
    own = [a for _, areas in race["communities"] for a in areas]

    # Decorative title row - column B is empty, so the widget ignores it.
    ws.cell(row=row, column=1, value=race["title"]).font = Font(bold=True, size=12, color=INK)
    for c in range(1, len(header) + 1):
        ws.cell(row=row, column=c).fill = TITLE_FILL
    ws.cell(row=row, column=3,
            value="White cells only: " + ", ".join(lbl for lbl, _ in race["communities"])
            ).font = Font(italic=True, size=10, color="6A6660")
    ws.row_dimensions[row].height = 20
    row += 1

    for name, home in race["cands"] + [("Write-in", ""), ("Blanks", "")]:
        ws.cell(row=row, column=1, value=race["key"]).font = Font(size=9, color="8A867F")
        item = ws.cell(row=row, column=2, value=name)
        item.font = Font(bold=name in ("Write-in", "Blanks"), size=11)
        if name in ("Write-in", "Blanks"):
            item.fill = SUB_FILL
        for area in COLUMNS:
            cell = ws.cell(row=row, column=col_of[area])
            cell.border = BOX
            cell.alignment = Alignment(horizontal="center")
            if area in own:
                cell.fill = ENTRY_FILL
                cell.protection = Protection(locked=False)
                dv.add(cell)
            else:
                # Belongs to the other district. Shaded and locked: the page would
                # ignore a number here, but a reporter would think it had counted.
                cell.fill = NA_FILL
                cell.protection = Protection(locked=True)
        ws.cell(row=row, column=1).border = BOX
        ws.cell(row=row, column=2).border = BOX
        row += 1
    row += 1  # spacer

ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 24
for i in range(3, 3 + len(COLUMNS)):
    ws.column_dimensions[get_column_letter(i)].width = 14

ws.protection.sheet = True
ws.protection.formatCells = False

# ─────────────────────────────────────────────────────────────────────────────
# HOW TO USE
# ─────────────────────────────────────────────────────────────────────────────
how = wb.create_sheet("How to use")
lines = [
    ("Election night data entry", "h1"),
    ("Winchester News - Massachusetts state primary, Tuesday, Sept. 1, 2026", "sub"),
    ("", ""),
    ("The short version", "h2"),
    ("Type each community's totals into the Results tab as that community posts them. "
     "Winchester goes in precinct by precinct, from our own reporters. The live page picks "
     "everything up within 15 seconds. There is nothing to publish and nothing to press.", ""),
    ("", ""),
    ("Rules that matter", "h2"),
    ("1. Leave a cell BLANK until that community has actually reported. A blank means "
     "'not yet'. A zero means 'reported, and the count was zero'. The page treats these "
     "differently and shows a dash where nothing has been entered.", ""),
    ("2. Enter a community's whole column at once - every candidate, plus write-in and "
     "blanks if you have them. A community only counts as reporting once every CANDIDATE "
     "cell in its column has a number, and its votes are left out of the standings until "
     "then. A half-typed column shows as 'partial' so you can see it is mid-entry.", ""),
    ("3. Only the white cells belong to a race. The two districts do not overlap: a number "
     "in a grey cell belongs to the other district and is ignored. The grey cells are locked.", ""),
    ("4. Do not rename anything in column A or column B, and do not reorder or delete rows. "
     "Column A is the key the page matches on. If a name is misspelled there, that race "
     "stops appearing.", ""),
    ("5. Do not insert or delete columns. The column order is fixed and must match the page.", ""),
    ("6. Numbers only. No percent signs, no notes in the cell. Commas are fine.", ""),
    ("", ""),
    ("Where the numbers come from", "h2"),
    ("Winchester - our reporters, precinct by precinct, eight precincts split between the "
     "two districts (4, 5, 6, 7 in the 2nd; 1, 2, 3, 8 in the 5th).", ""),
    ("Every other community - that city or town's own posted results, entered as one total "
     "per candidate. Somerville, Medford and the Cambridge portion for the 2nd Middlesex; "
     "Malden, Melrose, Reading, Stoneham and Wakefield for the 5th.", ""),
    ("Cambridge is in the 2nd Middlesex only for WARDS 10 and 11 plus Ward 7 Precinct 1 "
     "and Ward 8 Precinct 1 - eleven reporting units in all. Ward 9 is NOT in this district. "
     "Do not enter a citywide Cambridge total: it would count voters who are not in this "
     "district.", ""),
    ("", ""),
    ("What the page will and will not say", "h2"),
    ("It shows who is LEADING among the communities that have fully reported, and which "
     "communities are still out.", ""),
    ("It does not declare a winner, even at 100 percent. These totals are compiled by hand "
     "and are unofficial. A call belongs in the story, from a reporter. (If the desk decides "
     "otherwise, set ALLOW_CALL to true in index.html.)", ""),
    ("", ""),
    ("If something looks wrong", "h2"),
    ("The page keeps the last good numbers on screen if the sheet is briefly unreachable, "
     "and the status line reads 'Results updating...' instead of 'Live'. It recovers by itself.", ""),
    ("If a number on the page is wrong, fix it here. The page recomputes from scratch every "
     "refresh, so a correction here corrects the page within 15 seconds.", ""),
]
r = 1
for text, kind in lines:
    c = how.cell(row=r, column=1, value=text)
    if kind == "h1":
        c.font = Font(bold=True, size=18, color=INK)
    elif kind == "h2":
        c.font = Font(bold=True, size=13, color=ACCENT)
    elif kind == "sub":
        c.font = Font(italic=True, size=11, color="6A6660")
    else:
        c.font = Font(size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        how.row_dimensions[r].height = max(15, 14 * (len(text) // 95 + 1))
    r += 1
how.column_dimensions["A"].width = 105

# ─────────────────────────────────────────────────────────────────────────────
# DISTRICT COMPOSITION
# ─────────────────────────────────────────────────────────────────────────────
ctx = wb.create_sheet("District composition")
ctx.append(["Race", "Community", "Sheet column(s)", "Reporting unit", "Precincts"])
for c in range(1, 6):
    cell = ctx.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEAD_FILL
r = 2
for race in RACES:
    for label, areas in race["communities"]:
        ctx.cell(row=r, column=1, value=race["title"])
        ctx.cell(row=r, column=2, value=label)
        ctx.cell(row=r, column=3, value=", ".join(areas))
        ctx.cell(row=r, column=4,
                 value="4 precincts, entered separately" if label.startswith("Winchester")
                 else "one town-level total")
        ctx.cell(row=r, column=5, value=PRECINCTS.get(label, "TO ESTABLISH"))
        for c in range(1, 6):
            ctx.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
r += 1
ctx.cell(row=r, column=1, value="Entry is by community: only Winchester is typed precinct by "
         "precinct. The Precincts column is the denominator for the page's secondary "
         "\"N of M precincts\" line - when a city posts its citywide total, all of its precincts "
         "count at once. The 5th Middlesex totals 67. The 2nd Middlesex shows no precinct line "
         "until Cambridge's ward composition is settled.").font = Font(italic=True, size=10)
r += 2
ctx.cell(row=r, column=1, value="Candidates as configured").font = Font(bold=True, size=12)
r += 1
for h, col in (("Race", 1), ("Candidate", 2), ("Home", 3)):
    ctx.cell(row=r, column=col, value=h).font = Font(bold=True)
r += 1
for race in RACES:
    for name, home in race["cands"]:
        ctx.cell(row=r, column=1, value=race["title"])
        ctx.cell(row=r, column=2, value=name)
        ctx.cell(row=r, column=3, value=home)
        r += 1
r += 1
ctx.cell(row=r, column=1,
         value="VERIFY every name, spelling and ballot order against the official ballot "
               "before polls close.").font = Font(bold=True, color="A03020")
for col, w in (("A", 42), ("B", 30), ("C", 44), ("D", 30), ("E", 11)):
    ctx.column_dimensions[col].width = w

OUT = Path(__file__).resolve().parent.parent / "data" / "middlesex-senate-primary-2026-results.xlsx"
OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print("wrote", OUT.relative_to(Path.cwd()) if OUT.is_relative_to(Path.cwd()) else OUT,
      "with", len(COLUMNS), "data columns")
